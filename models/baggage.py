import openpyxl
from models.passenger import Passenger
import uuid


class Baggage:
    def __init__(self, capacity):
        self.capacity = capacity
        self.passengers = []
        self.saved_ids = set()

    def add_passenger(self, passenger):
        if len(self.passengers) < self.capacity:
            self.passengers.append(passenger)
            return True
        return False

    def save_to_file(self, filename):
        try:
            try:
                wb = openpyxl.load_workbook(filename)
                ws = wb.active
            except FileNotFoundError:
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "Информация о багажах пассажиров"
                headers = [
                    "Номер рейса",
                    "Дата и время вылета",
                    "Пункт назначения",
                    "Фамилия пассажира",
                    "Количество мест багажа",
                    "Суммарный вес багажа",
                ]
                ws.append(headers)

            for passenger in self.passengers:

                if passenger.id is None:
                    passenger.id = str(uuid.uuid4())
                    self.saved_ids.add(passenger.id)
                    ws.append(
                        [
                            passenger.flight_number,
                            passenger.departure_datetime,
                            passenger.destination,
                            passenger.passenger_name,
                            passenger.count_baggage,
                            passenger.weight_baggage,
                        ]
                    )

            wb.save(filename)
            return True
        except Exception as e:
            print(f"Ошибка при сохранении в файл: {e}")
            return False

    def load_from_file(self, filename):
        try:
            wb = openpyxl.load_workbook(filename)
            ws = wb.active

            records_found = False
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row and len(row) >= 6 and all(cell is not None for cell in row):
                    records_found = True
                    (
                        flight_number,
                        departure_datetime,
                        destination,
                        passenger_name,
                        count_baggage,
                        weight_baggage,
                    ) = row
                    passenger = Passenger(
                        flight_number,
                        departure_datetime,
                        destination,
                        passenger_name,
                        count_baggage,
                        weight_baggage,
                    )
                    if len(self.passengers) < self.capacity:
                        passenger.id = str(uuid.uuid4())
                        self.saved_ids.add(passenger.id)
                        self.add_passenger(passenger)
                    else:
                        raise Exception(
                            "Были выгружены не все пассажиры, недостаточно вместимости багажа."
                        )
            if not records_found:
                raise ValueError(f"В файле {filename} нет записей!")

        except FileNotFoundError:
            raise FileNotFoundError(f"Файл {filename} не найден.")

    def sort_by_name(self):
        self.passengers.sort(key=lambda passenger: passenger.passenger_name)
